#!/usr/bin/env python3
"""
Excel Processor
Handles comprehensive Excel file generation and Google Drive upload
Migrated from full-rounded-url-download-transcription.py
"""

import asyncio
import os
import sys
import time
from datetime import datetime
from typing import List, Dict, Any, Optional

# Add project root to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from core.processors.base_processor import BaseProcessor
from system.database import db_manager
from system.config import settings
from system.error_recovery import retry_async, RetryConfig, GOOGLE_API_RETRY_CONFIG

# Excel libraries
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.exceptions import InvalidFileException

# Google Drive API
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload


class ExcelProcessor(BaseProcessor):
    """Handles comprehensive Excel file generation and Google Drive upload"""
    
    def __init__(self):
        super().__init__("ExcelProcessor")
        self.processed_count = 0
        self.failed_count = 0
        
        # Configuration
        self.transcripts_dir = "assets/downloads/transcripts"
        self.excel_filename = os.getenv("EXCEL_FILENAME", "video_transcripts.xlsx")
        self.excel_file_path = os.path.join(self.transcripts_dir, self.excel_filename)
        # Use a specific folder for Excel files, not the general GOOGLE_DRIVE_FOLDER env var
        self.drive_folder = "VideoTranscripts"
        self.log_step(f"Excel processor initialized with drive folder: {self.drive_folder}")
        
        # Google Drive configuration
        self.scopes = ['https://www.googleapis.com/auth/drive.file']
        self.credentials_file = settings.google_credentials_file
        self.token_file = settings.google_token_file
        
        # Excel columns definition - matching the old implementation exactly
        self.columns = [
            # Basic Info
            "Index", "Generated Name", "Original Title", "Description", "Date Processed",
            
            # Creator Info  
            "Username", "Uploader ID", "Channel ID", "Channel URL",
            
            # Video Details
            "Video ID", "Platform", "Duration (seconds)", "Resolution", "FPS", "Format",
            
            # Engagement Metrics
            "View Count", "Like Count", "Comment Count", "Upload Date",
            
            # File Information
            "Video File Size (MB)", "Video Path", "Thumbnail Path", "Transcript Path", "Audio Path",
            
            # Content
            "Transcript", "Transcript Word Count",
            
            # Processing Info
            "Source URL", "Status", "Processing Time (seconds)", "Notes", "Error Details"
        ]
    
    async def initialize(self) -> bool:
        """Initialize Excel processor"""
        try:
            self.log_step("Initializing Excel processor")
            
            # Create transcripts directory
            os.makedirs(self.transcripts_dir, exist_ok=True)
            
            self.initialized = True
            self.status = "ready"
            self.log_step("Excel processor initialized successfully")
            return True
            
        except Exception as e:
            self.log_error("Failed to initialize Excel processor", e)
            return False
    
    async def process(self, urls: List[str] = None) -> bool:
        """Main processing method - generate and upload Excel file"""
        return await self.generate_and_upload_excel()
    
    async def generate_and_upload_excel(self) -> bool:
        """Generate comprehensive Excel file and upload to Google Drive"""
        try:
            self.log_step("Starting Excel file generation and upload")
            self.status = "processing"
            
            # Get all videos and thumbnails from database
            videos = await db_manager.get_all_videos()
            thumbnails = await db_manager.get_all_thumbnails()
            
            if not videos:
                self.log_step("No videos found to include in Excel file")
                return True
            
            # Generate Excel file
            excel_path = await self._generate_excel_file(videos, thumbnails)
            if not excel_path:
                self.log_error("Failed to generate Excel file")
                return False
            
            # Upload to Google Drive
            upload_success = await self._upload_excel_to_drive(excel_path)
            if upload_success:
                self.processed_count = len(videos)
                self.status = "completed"
                self.log_step(f"Excel file generated and uploaded successfully with {self.processed_count} entries")
                return True
            else:
                self.failed_count = 1
                self.status = "error"
                self.log_error("Failed to upload Excel file to Google Drive")
                return False
            
        except Exception as e:
            self.log_error("Error in Excel generation and upload", e)
            self.status = "error"
            return False
    
    async def _generate_excel_file(self, videos: List[Dict], thumbnails: List[Dict]) -> Optional[str]:
        """Generate comprehensive Excel file with video data"""
        try:
            self.log_step(f"Generating Excel file with {len(videos)} videos")
            
            # Create or load workbook
            wb, ws = await self._get_or_create_workbook()
            
            # Process each video
            for index, video in enumerate(videos, 1):
                # Find matching thumbnail
                video_filename = video.get('filename', '')
                base_name = os.path.splitext(video_filename)[0]
                matching_thumbnail = None
                
                for thumbnail in thumbnails:
                    if base_name in thumbnail.get('filename', '') or base_name in thumbnail.get('video_filename', ''):
                        matching_thumbnail = thumbnail
                        break
                
                # Prepare video data
                video_data = await self._prepare_video_data(video, matching_thumbnail, index)
                
                # Add row to Excel
                await self._add_video_row(ws, video_data)
            
            # Save workbook
            wb.save(self.excel_file_path)
            self.log_step(f"Excel file saved: {self.excel_file_path}")
            
            return self.excel_file_path
            
        except Exception as e:
            self.log_error("Error generating Excel file", e)
            return None
    
    async def _get_or_create_workbook(self) -> tuple:
        """Get existing workbook or create new one"""
        try:
            if os.path.exists(self.excel_file_path):
                try:
                    wb = load_workbook(self.excel_file_path)
                    ws = wb.active
                    self.log_step("Loaded existing Excel file")
                    return wb, ws
                except InvalidFileException:
                    # Backup corrupted file
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    backup_path = f"{self.excel_file_path}.backup_{timestamp}"
                    os.rename(self.excel_file_path, backup_path)
                    self.log_step(f"Moved corrupted Excel file to: {backup_path}")
                    
                    wb = Workbook()
                    ws = wb.active
                    await self._setup_workbook(wb, ws)
                    return wb, ws
            else:
                wb = Workbook()
                ws = wb.active
                await self._setup_workbook(wb, ws)
                return wb, ws
                
        except Exception as e:
            self.log_error("Error getting/creating workbook", e)
            raise
    
    async def _setup_workbook(self, wb: Workbook, ws) -> None:
        """Setup workbook with headers and validation"""
        try:
            # Add headers
            ws.append(self.columns)
            
            # Add data validation for Status column
            status_validation = DataValidation(
                type="list", 
                formula1='"In Progress,Completed,Failed,Needs Review,Skipped"', 
                allow_blank=False
            )
            ws.add_data_validation(status_validation)
            status_validation.add("V2:V1048576")  # Status column
            
            self.log_step("Workbook setup completed with headers and validation")
            
        except Exception as e:
            self.log_error("Error setting up workbook", e)
            raise
    
    async def _prepare_video_data(self, video: Dict, thumbnail: Optional[Dict], index: int) -> Dict[str, Any]:
        """Prepare comprehensive video data for Excel row"""
        try:
            # Get transcript text
            transcript_text = video.get('transcription_text', '')
            word_count = len(transcript_text.split()) if transcript_text else 0
            
            # Calculate resolution
            resolution = ""
            if video.get('width') and video.get('height'):
                resolution = f"{video['width']}x{video['height']}"
            
            # Get file size in MB
            file_size_mb = 0
            if video.get('file_path') and os.path.exists(video.get('file_path', '')):
                file_size_mb = os.path.getsize(video.get('file_path', '')) / (1024 * 1024)
            
            # Prepare comprehensive data with all metadata
            video_data = {
                'index': index,
                'generated_name': video.get('smart_name', ''),
                'title': video.get('title', ''),
                'description': video.get('description', ''),
                'date_processed': video.get('created_at', ''),
                'username': video.get('username', ''),
                'uploader_id': video.get('uploader_id', ''),
                'channel_id': video.get('channel_id', ''),
                'channel_url': video.get('channel_url', ''),
                'video_id': video.get('video_id', ''),
                'platform': video.get('platform', ''),
                'duration': video.get('duration', 0),
                'resolution': resolution,
                'fps': video.get('fps', ''),
                'format': video.get('format_id', ''),
                'view_count': video.get('view_count', ''),
                'like_count': video.get('like_count', ''),
                'comment_count': video.get('comment_count', ''),
                'upload_date': video.get('upload_date', ''),
                'file_size_mb': f"{file_size_mb:.2f}",
                'video_path': video.get('file_path', ''),
                'thumbnail_path': thumbnail.get('file_path', '') if thumbnail else '',
                'transcript_path': f"{video.get('smart_name', '')}.txt" if video.get('smart_name') else '',
                'audio_path': '',  # Not stored in database
                'transcript': transcript_text,
                'word_count': word_count,
                'source_url': video.get('webpage_url', video.get('url', '')),
                'status': 'Completed' if video.get('transcription_status') == 'COMPLETED' else 'Pending',
                'processing_time': 0,  # Not tracked in database
                'notes': '',
                'error_details': ''
            }
            
            return video_data
            
        except Exception as e:
            self.log_error("Error preparing video data", e)
            return {}
    
    async def _add_video_row(self, ws, video_data: Dict[str, Any]) -> None:
        """Add video data as row to worksheet"""
        try:
            row_data = [
                video_data.get('index', ''),
                video_data.get('generated_name', ''),
                video_data.get('title', ''),
                video_data.get('description', ''),
                video_data.get('date_processed', ''),
                video_data.get('username', ''),
                video_data.get('uploader_id', ''),
                video_data.get('channel_id', ''),
                video_data.get('channel_url', ''),
                video_data.get('video_id', ''),
                video_data.get('platform', ''),
                video_data.get('duration', 0),
                video_data.get('resolution', ''),
                video_data.get('fps', ''),
                video_data.get('format', ''),
                video_data.get('view_count', ''),
                video_data.get('like_count', ''),
                video_data.get('comment_count', ''),
                video_data.get('upload_date', ''),
                video_data.get('file_size_mb', ''),
                video_data.get('video_path', ''),
                video_data.get('thumbnail_path', ''),
                video_data.get('transcript_path', ''),
                video_data.get('audio_path', ''),
                video_data.get('transcript', ''),  # Full transcript
                video_data.get('word_count', 0),
                video_data.get('source_url', ''),
                video_data.get('status', 'In Progress'),
                video_data.get('processing_time', 0),
                video_data.get('notes', ''),
                video_data.get('error_details', '')
            ]
            
            ws.append(row_data)
            
            # Apply validation to new row
            new_row = ws.max_row
            # Note: Data validation is already applied to the entire column range
            # No need to add validation to individual rows
            
        except Exception as e:
            self.log_error("Error adding video row", e)
    
    @retry_async(GOOGLE_API_RETRY_CONFIG)
    async def _upload_excel_to_drive(self, excel_path: str) -> bool:
        """Upload Excel file to Google Drive with retry logic"""
        try:
            self.log_step(f"Uploading Excel file to Google Drive: {os.path.basename(excel_path)}")
            
            service = await self._get_drive_service()
            if not service:
                return False
            
            # Create or find folder
            folder_id = await self._get_or_create_drive_folder(service)
            if not folder_id:
                return False
            
            # Upload file
            file_id = await self._upload_file_to_drive(service, excel_path, folder_id)
            if file_id:
                self.log_step(f"Successfully uploaded Excel file to Google Drive")
                return True
            else:
                self.log_error("Failed to upload Excel file to Google Drive")
                return False
                
        except Exception as e:
            self.log_error("Error uploading Excel to Drive", e)
            return False
    
    async def _get_drive_service(self) -> Optional[Any]:
        """Get authenticated Google Drive service"""
        try:
            creds = None
            if os.path.exists(self.token_file):
                try:
                    creds = Credentials.from_authorized_user_file(self.token_file, self.scopes)
                    self.log_step("Loaded existing credentials from token file")
                except Exception as e:
                    self.log_error(f"Error reading token file: {str(e)}")
                    creds = None
                    
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    try:
                        creds.refresh(Request())
                        with open(self.token_file, 'w') as token:
                            token.write(creds.to_json())
                        self.log_step("Successfully refreshed credentials")
                    except Exception as e:
                        self.log_error(f"Error refreshing credentials: {str(e)}")
                        creds = None
                
                # If we still don't have valid creds, start fresh OAuth flow
                if not creds:
                    try:
                        self.log_step("Starting OAuth flow for new authentication")
                        flow = InstalledAppFlow.from_client_secrets_file(self.credentials_file, self.scopes)
                        creds = flow.run_local_server(
                            port=8080,
                            access_type='offline',
                            prompt='consent'
                        )
                        # Save the new credentials
                        with open(self.token_file, 'w') as token:
                            token.write(creds.to_json())
                        self.log_step("New authentication tokens obtained and saved")
                    except Exception as e:
                        self.log_error(f"Error in OAuth flow: {str(e)}")
                        return None
            
            service = build('drive', 'v3', credentials=creds)
            self.log_step("Google Drive service initialized successfully")
            return service
            
        except Exception as e:
            self.log_error(f"Failed to initialize Google Drive service: {str(e)}")
            return None
    
    async def _get_or_create_drive_folder(self, service) -> Optional[str]:
        """Get or create Google Drive folder"""
        try:
            # Search for existing folder
            folder_query = f"name='{self.drive_folder}' and mimeType='application/vnd.google-apps.folder'"
            self.log_step(f"Searching for folder: {self.drive_folder}")
            results = service.files().list(q=folder_query).execute()
            folders = results.get('files', [])
            
            # Debug: List all found folders
            if folders:
                for folder in folders:
                    self.log_step(f"Found folder: {folder.get('name')} (ID: {folder.get('id')})")
            
            if folders:
                folder_id = folders[0]['id']
                self.log_step(f"Using existing folder: {self.drive_folder} (ID: {folder_id})")
                return folder_id
            else:
                # Create new folder
                folder_metadata = {'name': self.drive_folder, 'mimeType': 'application/vnd.google-apps.folder'}
                folder = service.files().create(body=folder_metadata).execute()
                folder_id = folder['id']
                self.log_step(f"Created new folder: {self.drive_folder}")
                return folder_id
                
        except Exception as e:
            self.log_error("Error getting/creating Drive folder", e)
            return None
    
    async def _upload_file_to_drive(self, service, file_path: str, folder_id: str) -> Optional[str]:
        """Upload file to Google Drive"""
        try:
            filename = os.path.basename(file_path)
            file_size = os.path.getsize(file_path)
            
            self.log_step(f"Uploading file: {filename} ({file_size / (1024*1024):.2f} MB)")
            
            # Check if file already exists
            file_query = f"name='{filename}' and '{folder_id}' in parents"
            existing = service.files().list(q=file_query).execute().get('files', [])
            
            media = MediaFileUpload(file_path, resumable=True)
            
            if existing:
                # Update existing file
                file_id = existing[0]['id']
                service.files().update(fileId=file_id, media_body=media).execute()
                self.log_step(f"Updated existing file: {filename}")
                return file_id
            else:
                # Create new file
                file_metadata = {'name': filename, 'parents': [folder_id]}
                file = service.files().create(body=file_metadata, media_body=media).execute()
                file_id = file.get('id')
                self.log_step(f"Created new file: {filename}")
                return file_id
                
        except Exception as e:
            self.log_error(f"Error uploading file to Drive: {str(e)}")
            return None
    
    async def cleanup(self) -> None:
        """Cleanup Excel processor resources"""
        try:
            self.log_step("Cleaning up Excel processor")
            self.status = "idle"
            self.log_step("Excel processor cleanup completed")
        except Exception as e:
            self.log_error("Error during Excel processor cleanup", e)
