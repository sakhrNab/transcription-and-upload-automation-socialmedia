#!/usr/bin/env python3
"""
Enhanced Video Downloader with Transcription & Comprehensive Logging
Combines yt_dlp reliability with intensive logging and proper file organization.

Features:
- Downloads videos using yt_dlp (reliable) 
- Downloads thumbnails when available
- Converts video to audio and transcribes using Whisper
- Generates smart video names using GPT-4o-mini
- Saves transcripts as separate text files
- Comprehensive Excel logging with all metadata
- Intensive session logging with performance metrics
- Google Drive upload
- Proper file organization in separate folders
- Batch processing support
"""

import os
import re
import json
import time
import argparse
import concurrent.futures
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Dict, Any, List

# Core libraries
import yt_dlp
import ffmpeg
import whisper
import requests
import torch
import gc
from openai import OpenAI
from dotenv import load_dotenv

# Excel and Google Drive
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.exceptions import InvalidFileException
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.auth.exceptions import RefreshError
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# Load environment variables
load_dotenv()

# ---------------------------
# Configuration from .env
# ---------------------------
# Directory structure
VIDEO_OUTPUT_DIR = os.getenv("VIDEO_OUTPUT_DIR", "downloads/videos")
AUDIO_OUTPUT_DIR = os.getenv("AUDIO_OUTPUT_DIR", "downloads/audio") 
THUMBNAILS_DIR = os.getenv("THUMBNAILS_DIR", "downloads/thumbnails")
TRANSCRIPTS_DIR = os.getenv("TRANSCRIPTS_DIR", "downloads/transcripts")
LOGS_DIR = os.getenv("LOGS_DIR", "logs")

# Excel configuration
EXCEL_FILENAME = os.getenv("EXCEL_FILENAME", "video_transcripts.xlsx")
EXCEL_FILE_PATH = os.getenv("EXCEL_FILE_PATH", os.path.join(TRANSCRIPTS_DIR, EXCEL_FILENAME))

# Processing configuration
WHISPER_MODEL = os.getenv("WHISPER_MODEL", "base") # tiny, base, small, medium, large
MAX_AUDIO_DURATION = int(os.getenv("MAX_AUDIO_DURATION", "1800")) # 30 minutes
CHUNK_DURATION = int(os.getenv("CHUNK_DURATION", "30")) # seconds
KEEP_AUDIO_FILES = os.getenv("KEEP_AUDIO_FILES", "true").lower() == "true"
KEEP_TEMP_FILES = os.getenv("KEEP_TEMP_FILES", "false").lower() == "true"

# Google Drive configuration
SCOPES = ['https://www.googleapis.com/auth/drive.file']
CREDENTIALS_FILE = os.getenv("GOOGLE_CREDENTIALS_FILE", "credentials.json")
TOKEN_FILE = os.getenv("GOOGLE_TOKEN_FILE", "token.json")
DRIVE_FOLDER = os.getenv("GOOGLE_DRIVE_FOLDER", "VideoTranscripts")

# OpenAI configuration
api_key = os.getenv("OPENAI_API_KEY")
if not api_key:
    raise ValueError("OPENAI_API_KEY not set. Please add it to your .env file.")
client = OpenAI(api_key=api_key)

# Create all directories
for directory in [VIDEO_OUTPUT_DIR, AUDIO_OUTPUT_DIR, THUMBNAILS_DIR, TRANSCRIPTS_DIR, LOGS_DIR]:
    os.makedirs(directory, exist_ok=True)

# ---------------------------
# Intensive Logging System
# ---------------------------
class SessionLogger:
    def __init__(self):
        self.session_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.session_start = datetime.now(timezone.utc)
        self.logs_dir = Path(LOGS_DIR)
        self.logs_dir.mkdir(exist_ok=True)
        
        # Initialize log files
        self.session_log_file = self.logs_dir / f"session_{self.session_id}.json"
        self.debug_log_file = self.logs_dir / f"debug_{self.session_id}.json"
        self.error_log_file = self.logs_dir / f"errors_{self.session_id}.json"
        self.performance_log_file = self.logs_dir / "performance_history.json"
        
        self.session_data = {
            'session_id': self.session_id,
            'start_time': self.session_start.isoformat(),
            'config': {
                'video_dir': VIDEO_OUTPUT_DIR,
                'audio_dir': AUDIO_OUTPUT_DIR,
                'thumbnails_dir': THUMBNAILS_DIR,
                'transcripts_dir': TRANSCRIPTS_DIR,
                'whisper_model': WHISPER_MODEL,
                'excel_path': EXCEL_FILE_PATH
            },
            'processing_log': [],
            'errors': [],
            'performance_metrics': {},
            'file_operations': [],
            'summary': {}
        }
    
    def log_step(self, video_index: int, step: str, details: Dict[str, Any]):
        """Log a processing step with timestamp and details"""
        entry = {
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'video_index': video_index,
            'step': step,
            'details': details
        }
        self.session_data['processing_log'].append(entry)
        print(f"[{video_index}] {step}: {details.get('message', '')}")
    
    def log_error(self, video_index: int, error: str, context: Dict[str, Any] = None):
        """Log an error with context"""
        error_entry = {
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'video_index': video_index,
            'error': error,
            'context': context or {}
        }
        self.session_data['errors'].append(error_entry)
        self._write_error_log(error_entry)
        print(f"[{video_index}] ERROR: {error}")
    
    def log_file_operation(self, operation: str, source: str, destination: str, success: bool, size_bytes: int = 0):
        """Log file operations (downloads, conversions, moves)"""
        entry = {
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'operation': operation,
            'source': source,
            'destination': destination,
            'success': success,
            'size_bytes': size_bytes,
            'size_mb': round(size_bytes / (1024*1024), 2) if size_bytes > 0 else 0
        }
        self.session_data['file_operations'].append(entry)
    
    def finalize_session(self, results: List[Dict[str, Any]]):
        """Finalize session with comprehensive summary and performance metrics"""
        session_end = datetime.now(timezone.utc)
        duration = (session_end - self.session_start).total_seconds()
        
        # Calculate metrics
        successful = len([r for r in results if r['status'] == 'Completed'])
        failed = len([r for r in results if r['status'] == 'Failed'])
        warnings = len([r for r in results if r['status'] == 'Needs Review'])
        skipped = len([r for r in results if r['status'] == 'Skipped'])
        
        total_size = sum([r.get('filesize', 0) for r in results])
        avg_processing_time = sum([r.get('processing_time', 0) for r in results]) / len(results) if results else 0
        
        self.session_data.update({
            'end_time': session_end.isoformat(),
            'duration_seconds': duration,
            'summary': {
                'total_videos': len(results),
                'successful': successful,
                'failed': failed,
                'warnings': warnings,
                'skipped': skipped,
                'success_rate': round((successful / len(results)) * 100, 1) if results else 0,
                'total_download_size_mb': round(total_size / (1024*1024), 2),
                'avg_processing_time_seconds': round(avg_processing_time, 2),
                'videos_per_minute': round((len(results) / duration) * 60, 2) if duration > 0 else 0
            },
            'performance_metrics': {
                'download_efficiency': round((successful / len(results)) * 100, 1) if results else 0,
                'error_rate': round((failed / len(results)) * 100, 1) if results else 0,
                'data_throughput_mbps': round((total_size / duration) / (1024*1024), 2) if duration > 0 else 0
            }
        })
        
        # Write session log
        self._write_session_log()
        
        # Update performance history
        self._update_performance_history()
        
        # Print summary
        self._print_session_summary()
    
    def _write_session_log(self):
        """Write comprehensive session log"""
        try:
            self.session_log_file.write_text(
                json.dumps(self.session_data, indent=2, ensure_ascii=False, default=str),
                encoding='utf-8'
            )
        except Exception as e:
            print(f"Failed to write session log: {e}")
    
    def _write_error_log(self, error_entry: Dict[str, Any]):
        """Write individual error to error log"""
        try:
            if self.error_log_file.exists():
                existing = json.loads(self.error_log_file.read_text(encoding='utf-8'))
                if isinstance(existing, list):
                    existing.append(error_entry)
                else:
                    existing = [error_entry]
            else:
                existing = [error_entry]
            
            self.error_log_file.write_text(
                json.dumps(existing, indent=2, ensure_ascii=False, default=str),
                encoding='utf-8'
            )
        except Exception as e:
            print(f"Failed to write error log: {e}")
    
    def _update_performance_history(self):
        """Update performance history for tracking trends"""
        try:
            history = []
            if self.performance_log_file.exists():
                history = json.loads(self.performance_log_file.read_text(encoding='utf-8'))
            
            history.append({
                'session_id': self.session_id,
                'timestamp': self.session_data['end_time'],
                'metrics': self.session_data['performance_metrics'],
                'summary': self.session_data['summary']
            })
            
            # Keep only last 50 sessions
            history = history[-50:]
            
            self.performance_log_file.write_text(
                json.dumps(history, indent=2, ensure_ascii=False, default=str),
                encoding='utf-8'
            )
        except Exception as e:
            print(f"Failed to update performance history: {e}")
    
    def _print_session_summary(self):
        """Print detailed session summary"""
        summary = self.session_data['summary']
        metrics = self.session_data['performance_metrics']
        
        print(f"\n{'='*80}")
        print(f"SESSION SUMMARY - {self.session_id}")
        print(f"{'='*80}")
        print(f"Duration: {self.session_data['duration_seconds']:.1f}s ({self.session_data['duration_seconds']/60:.1f} minutes)")
        print(f"Videos processed: {summary['total_videos']}")
        print(f"Success rate: {summary['success_rate']}%")
        print(f"Average time per video: {summary['avg_processing_time_seconds']}s")
        print(f"Total data downloaded: {summary['total_download_size_mb']} MB")
        print(f"Processing throughput: {summary['videos_per_minute']} videos/minute")
        print(f"Data throughput: {metrics['data_throughput_mbps']} MB/s")
        print(f"\nResults breakdown:")
        print(f" ✅ Successful: {summary['successful']}")
        print(f" ⚠️ Warnings: {summary['warnings']}")
        print(f" ❌ Failed: {summary['failed']}")
        print(f" ⏭️ Skipped: {summary['skipped']}")
        print(f"\nLogs saved to: {self.logs_dir}")
        print(f"{'='*80}")

# Global logger instance
logger = SessionLogger()

# ---------------------------
# Utility Functions
# ---------------------------
def get_unique_filename(path):
    """Generate unique filename if file exists"""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    counter = 1
    while True:
        new_path = f"{base}_{counter}{ext}"
        if not os.path.exists(new_path):
            return new_path
        counter += 1

def safe_filename(text: str, max_length: int = 100) -> str:
    """Convert text to safe filename"""
    if not text:
        return "unknown"
    safe = re.sub(r'[<>:"/\\|?*]', '_', text)
    safe = re.sub(r'\s+', '_', safe).strip('_')
    return safe[:max_length] if len(safe) > max_length else safe

def get_video_number(output_dir: str, username: str) -> int:
    """Get next available number for a username"""
    pattern = re.compile(rf'\d+_{re.escape(username)}_.*')
    max_num = 0
    
    if os.path.exists(output_dir):
        for filename in os.listdir(output_dir):
            if pattern.match(filename):
                try:
                    num = int(filename.split('_')[0])
                    max_num = max(max_num, num)
                except (ValueError, IndexError):
                    continue
    return max_num + 1

# ---------------------------
# Video Processing Functions
# ---------------------------
def download_video_and_metadata(url: str, index: int) -> tuple[str, dict, dict]:
    """Download video using yt_dlp and extract comprehensive metadata"""
    logger.log_step(index, "METADATA_EXTRACTION", {"message": "Extracting video information", "url": url})
    
    # First, extract info without downloading
    with yt_dlp.YoutubeDL({'quiet': True}) as ydl:
        try:
            info = ydl.extract_info(url, download=False)
            logger.log_step(index, "METADATA_SUCCESS", {
                "message": f"Extracted metadata for {info.get('title', 'Unknown')}",
                "platform": info.get('extractor', 'Unknown'),
                "duration": info.get('duration', 0)
            })
        except Exception as e:
            logger.log_error(index, f"Failed to extract video info: {str(e)}", {"url": url})
            raise Exception(f"Failed to extract video info: {str(e)}")
    
    # Extract and clean metadata
    video_id = info.get('id', 'unknown')
    username = info.get('uploader', info.get('channel', 'unknown'))
    username = re.sub(r'[^\w]+', '_', username.lower()) if username else 'unknown'
    title = info.get('title', video_id)
    description = info.get('description', '')
    
    # Check if already downloaded
    for file in os.listdir(VIDEO_OUTPUT_DIR):
        if video_id in file and file.endswith(('.mp4', '.webm', '.mkv')):
            full_path = os.path.join(VIDEO_OUTPUT_DIR, file)
            logger.log_step(index, "ALREADY_EXISTS", {"message": f"Video already downloaded: {file}"})
            
            metadata = extract_comprehensive_metadata(info, full_path)
            return full_path, metadata, info
    
    # Create filename with sequential numbering
    seq_num = get_video_number(VIDEO_OUTPUT_DIR, username)
    filename_template = os.path.join(VIDEO_OUTPUT_DIR, f"{seq_num:02d}_{username}_{video_id}.%(ext)s")
    
    logger.log_step(index, "DOWNLOAD_START", {
        "message": f"Starting download: {title}",
        "filename_template": filename_template,
        "expected_size": info.get('filesize_approx', 'Unknown')
    })
    
    # Download configuration
    ydl_opts = {
        'outtmpl': filename_template,
        'format': 'best[ext=mp4]/best',
        'writesubtitles': False, # We'll transcribe instead
        'writeautomaticsub': False,
        'ignoreerrors': False,
        'quiet': True
    }
    
    download_start = time.time()
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        ydl.download([url])
        downloaded_file = ydl.prepare_filename(info)
    
    download_time = time.time() - download_start
    
    if not os.path.exists(downloaded_file):
        # Try to find the downloaded file with different extension
        base_name = os.path.splitext(downloaded_file)[0]
        for ext in ['.mp4', '.webm', '.mkv', '.mov']:
            alt_path = base_name + ext
            if os.path.exists(alt_path):
                downloaded_file = alt_path
                break
        else:
            logger.log_error(index, "Downloaded file not found", {"expected_path": downloaded_file})
            raise Exception("Downloaded file not found")
    
    file_size = os.path.getsize(downloaded_file)
    logger.log_file_operation("DOWNLOAD", url, downloaded_file, True, file_size)
    logger.log_step(index, "DOWNLOAD_SUCCESS", {
        "message": f"Downloaded successfully",
        "filename": os.path.basename(downloaded_file),
        "size_mb": round(file_size / (1024*1024), 2),
        "download_time": round(download_time, 2)
    })
    
    metadata = extract_comprehensive_metadata(info, downloaded_file)
    return downloaded_file, metadata, info

def extract_comprehensive_metadata(info: dict, file_path: str) -> dict:
    """Extract all available metadata from yt_dlp info"""
    return {
        'video_id': info.get('id', 'unknown'),
        'username': re.sub(r'[^\w]+', '_', (info.get('uploader', '') or 'unknown').lower()),
        'title': info.get('title', ''),
        'description': info.get('description', ''),
        'duration': info.get('duration', 0),
        'view_count': info.get('view_count'),
        'like_count': info.get('like_count'),
        'comment_count': info.get('comment_count'),
        'upload_date': info.get('upload_date', ''),
        'uploader_id': info.get('uploader_id', ''),
        'uploader_url': info.get('uploader_url', ''),
        'channel_id': info.get('channel_id', ''),
        'channel_url': info.get('channel_url', ''),
        'thumbnail_url': info.get('thumbnail', ''),
        'webpage_url': info.get('webpage_url', ''),
        'extractor': info.get('extractor', ''),
        'platform': info.get('extractor_key', ''),
        'format_id': info.get('format_id', ''),
        'width': info.get('width'),
        'height': info.get('height'),
        'fps': info.get('fps'),
        'filesize': os.path.getsize(file_path) if os.path.exists(file_path) else 0,
        'file_path': file_path
    }

def download_thumbnail(thumbnail_url: str, video_id: str, username: str, index: int) -> Optional[str]:
    """Download thumbnail image to thumbnails directory"""
    if not thumbnail_url:
        logger.log_step(index, "THUMBNAIL_SKIP", {"message": "No thumbnail URL available"})
        return None
    
    try:
        logger.log_step(index, "THUMBNAIL_START", {"message": "Downloading thumbnail", "url": thumbnail_url})
        
        response = requests.get(thumbnail_url, timeout=15, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
        if response.status_code == 200:
            # Get sequential number for the username
            seq_num = get_video_number(THUMBNAILS_DIR, username)

            # Determine file extension from content type
            content_type = response.headers.get('content-type', '')
            if 'jpeg' in content_type or 'jpg' in content_type:
                ext = '.jpg'
            elif 'png' in content_type:
                ext = '.png'
            elif 'webp' in content_type:
                ext = '.webp'
            else:
                ext = '.jpg'
            
            filename = f"{seq_num:02d}_{username}_{video_id}{ext}"
            filepath = os.path.join(THUMBNAILS_DIR, filename)
            filepath = get_unique_filename(filepath)
            
            with open(filepath, 'wb') as f:
                f.write(response.content)
            
            logger.log_file_operation("THUMBNAIL_DOWNLOAD", thumbnail_url, filepath, True, len(response.content))
            logger.log_step(index, "THUMBNAIL_SUCCESS", {
                "message": f"Downloaded thumbnail: {os.path.basename(filepath)}",
                "size_kb": round(len(response.content) / 1024, 2)
            })
            return filepath
        else:
            logger.log_step(index, "THUMBNAIL_FAILED", {
                "message": f"HTTP {response.status_code}",
                "url": thumbnail_url
            })
            
    except Exception as e:
        logger.log_error(index, f"Thumbnail download failed: {str(e)}", {"url": thumbnail_url})
    
    return None

def convert_video_to_audio(video_file: str, index: int) -> str:
    """Convert video to high-quality audio for transcription"""
    video_basename = os.path.splitext(os.path.basename(video_file))[0]
    audio_file = os.path.join(AUDIO_OUTPUT_DIR, f"{video_basename}.wav")
    audio_file = get_unique_filename(audio_file)
    
    logger.log_step(index, "AUDIO_CONVERSION_START", {
        "message": "Converting video to audio",
        "input": os.path.basename(video_file),
        "output": os.path.basename(audio_file)
    })
    
    try:
        conversion_start = time.time()
        ffmpeg.input(video_file).output(
            audio_file,
            acodec='pcm_s16le',
            ac=1, # Mono
            ar='16000', # 16kHz sample rate (optimal for Whisper)
            loglevel='error'
        ).run(overwrite_output=True)
        
        conversion_time = time.time() - conversion_start
        audio_size = os.path.getsize(audio_file)
        
        logger.log_file_operation("AUDIO_CONVERSION", video_file, audio_file, True, audio_size)
        logger.log_step(index, "AUDIO_CONVERSION_SUCCESS", {
            "message": f"Audio conversion completed",
            "output_file": os.path.basename(audio_file),
            "size_mb": round(audio_size / (1024*1024), 2),
            "conversion_time": round(conversion_time, 2)
        })
        
        return audio_file
        
    except Exception as e:
        logger.log_error(index, f"Audio conversion failed: {str(e)}", {"video_file": video_file})
        raise Exception(f"Audio conversion failed: {str(e)}")

def transcribe_audio_with_whisper(audio_file: str, index: int) -> str:
    """Transcribe audio using Whisper with comprehensive logging"""
    logger.log_step(index, "TRANSCRIPTION_START", {
        "message": f"Starting transcription with {WHISPER_MODEL} model",
        "audio_file": os.path.basename(audio_file),
        "model": WHISPER_MODEL
    })
    
    try:
        # Clear GPU cache if available
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            logger.log_step(index, "GPU_CACHE_CLEARED", {"message": "Cleared CUDA cache"})
        
        transcription_start = time.time()
        
        # Load model
        model = whisper.load_model(WHISPER_MODEL)
        logger.log_step(index, "MODEL_LOADED", {"message": f"Loaded {WHISPER_MODEL} model"})
        
        # Check audio duration
        audio_duration = get_audio_duration(audio_file)
        if audio_duration > MAX_AUDIO_DURATION:
            logger.log_step(index, "LONG_AUDIO_DETECTED", {
                "message": f"Audio is {audio_duration}s, will chunk",
                "max_duration": MAX_AUDIO_DURATION
            })
            transcript = transcribe_long_audio(audio_file, model, index)
        else:
            # Standard transcription
            result = model.transcribe(audio_file, verbose=False)
            transcript = result['text'].strip()
        
        transcription_time = time.time() - transcription_start
        
        # Cleanup
        del model
        gc.collect()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
        
        if transcript:
            logger.log_step(index, "TRANSCRIPTION_SUCCESS", {
                "message": f"Transcription completed",
                "transcript_length": len(transcript),
                "word_count": len(transcript.split()),
                "transcription_time": round(transcription_time, 2)
            })
        else:
            logger.log_step(index, "TRANSCRIPTION_EMPTY", {"message": "Transcription produced empty result"})
        
        return transcript
        
    except Exception as e:
        logger.log_error(index, f"Transcription failed: {str(e)}", {"audio_file": audio_file})
        return ""

def get_audio_duration(audio_file: str) -> float:
    """Get audio duration using ffmpeg"""
    try:
        probe = ffmpeg.probe(audio_file)
        duration = float(probe['streams'][0]['duration'])
        return duration
    except Exception:
        return 0

def transcribe_long_audio(audio_file: str, model, index: int) -> str:
    """Handle long audio files by chunking"""
    logger.log_step(index, "CHUNKING_START", {"message": "Chunking long audio file"})
    
    try:
        # Use ffmpeg to split audio into chunks
        chunks = []
        duration = get_audio_duration(audio_file)
        num_chunks = int(duration // CHUNK_DURATION) + 1
        
        for i in range(num_chunks):
            start_time = i * CHUNK_DURATION
            chunk_file = os.path.join(AUDIO_OUTPUT_DIR, f"chunk_{i}_{os.path.basename(audio_file)}")
            
            ffmpeg.input(audio_file, ss=start_time, t=CHUNK_DURATION).output(
                chunk_file,
                acodec='pcm_s16le',
                ac=1,
                ar='16000',
                loglevel='error'
            ).run(overwrite_output=True)
            
            chunks.append(chunk_file)
        
        logger.log_step(index, "CHUNKS_CREATED", {"message": f"Created {len(chunks)} chunks"})
        
        # Transcribe each chunk
        full_transcript = []
        for i, chunk in enumerate(chunks):
            try:
                result = model.transcribe(chunk, verbose=False)
                chunk_transcript = result['text'].strip()
                full_transcript.append(chunk_transcript)
                logger.log_step(index, "CHUNK_TRANSCRIBED", {
                    "message": f"Transcribed chunk {i+1}/{len(chunks)}",
                    "chunk_length": len(chunk_transcript)
                })
            except Exception as e:
                logger.log_error(index, f"Failed to transcribe chunk {i+1}: {str(e)}")
            finally:
                # Clean up chunk file
                try:
                    os.remove(chunk)
                except Exception:
                    pass
        
        transcript = ' '.join(full_transcript)
        logger.log_step(index, "CHUNKING_COMPLETE", {
            "message": f"Completed chunked transcription",
            "total_length": len(transcript)
        })
        
        return transcript
        
    except Exception as e:
        logger.log_error(index, f"Chunked transcription failed: {str(e)}")
        return ""

def generate_smart_video_name(title: str, description: str, index: int) -> str:
    """Generate intelligent video name using GPT-4o-mini"""
    date_str = datetime.now().strftime('%d.%m.%Y')
    
    logger.log_step(index, "NAME_GENERATION_START", {"message": "Generating smart video name"})
    
    # Create context for naming
    context = f"Title: {title}\n\nDescription: {description[:500]}"
    
    prompt = f"""Based on this video content, generate a concise, descriptive name (1-3 words) that captures the main topic or tool mentioned:

{context}

Return only the name, no explanation. Make it suitable for a filename."""
    
    try:
        response = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=[
                {'role': 'system', 'content': 'You generate concise, descriptive names for video content.'},
                {'role': 'user', 'content': prompt}
            ],
            max_tokens=20,
            temperature=0.3
        )
        
        name = response.choices[0].message.content.strip()
        name = safe_filename(name.lower())
        generated_name = f"{name}_{date_str}"
        
        logger.log_step(index, "NAME_GENERATION_SUCCESS", {
            "message": f"Generated name: {generated_name}",
            "original_title": title[:50]
        })
        
        return generated_name
        
    except Exception as e:
        logger.log_error(index, f"Name generation failed: {str(e)}")
        # Fallback to title-based naming
        title_safe = safe_filename(title.lower())[:30]
        fallback_name = f"{title_safe}_{date_str}"
        
        logger.log_step(index, "NAME_GENERATION_FALLBACK", {
            "message": f"Using fallback name: {fallback_name}"
        })
        
        return fallback_name

def save_transcript_file(transcript: str, generated_name: str, metadata: dict, index: int) -> str:
    """Save transcript as separate text file with metadata header"""
    transcript_filename = f"{generated_name}.txt"
    transcript_path = os.path.join(TRANSCRIPTS_DIR, transcript_filename)
    transcript_path = get_unique_filename(transcript_path)
    
    # Create transcript file with metadata header
    header = f"""# Video Transcript
Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Video Title: {metadata.get('title', 'Unknown')}
Platform: {metadata.get('platform', 'Unknown')}
Duration: {metadata.get('duration', 0)} seconds
Video ID: {metadata.get('video_id', 'Unknown')}
Username: {metadata.get('username', 'Unknown')}
Source URL: {metadata.get('webpage_url', 'Unknown')}

{'='*50}
TRANSCRIPT:
{'='*50}

{transcript}
"""
    
    try:
        with open(transcript_path, 'w', encoding='utf-8') as f:
            f.write(header)
        
        logger.log_file_operation("TRANSCRIPT_SAVE", "memory", transcript_path, True, len(header.encode('utf-8')))
        logger.log_step(index, "TRANSCRIPT_SAVED", {
            "message": f"Saved transcript file: {os.path.basename(transcript_path)}",
            "transcript_length": len(transcript),
            "file_path": transcript_path
        })
        
        return transcript_path
        
    except Exception as e:
        logger.log_error(index, f"Failed to save transcript: {str(e)}")
        return ""

# ---------------------------
# Excel Management with Full Metadata
# ---------------------------
def update_excel_comprehensive(video_data: Dict[str, Any], excel_path: str):
    """Update Excel with comprehensive video data and metadata"""
    excel_dir = os.path.dirname(excel_path)
    if excel_dir:
        os.makedirs(excel_dir, exist_ok=True)
    
    # Define comprehensive columns
    columns = [
        # Basic Info
        "Index", "Generated Name", "Original Title", "Description", "Date Processed",
        
        # Creator Info  
        "Username", "Uploader ID", "Channel ID", "Channel URL",
        
        # Video Details
        "Video ID", "Platform", "Duration (seconds)", "Resolution",
        
        # Engagement Metrics
        "View Count", "Like Count", "Comment Count", "Upload Date",
        
        # File Information
        "Video File Size (MB)", "Video Path", "Thumbnail Path", "Transcript Path", "Audio Path",
        
        # Content
        "Transcript", "Transcript Word Count",
        
        # Processing Info
        "Source URL", "Status", "Processing Time (seconds)", "Notes", "Error Details"
    ]
    
    # Data validation for Status
    status_validation = DataValidation(
        type="list", 
        formula1='"In Progress,Completed,Failed,Needs Review,Skipped"', 
        allow_blank=False
    )
    
    # Load or create workbook
    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
        except InvalidFileException:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = f"{excel_path}.backup_{timestamp}"
            os.rename(excel_path, backup_path)
            print(f"Moved corrupted Excel file to: {backup_path}")
            
            wb = Workbook()
            ws = wb.active
            ws.append(columns)
            ws.add_data_validation(status_validation)
            status_validation.add("V2:V1048576") # Status column
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(columns)
        ws.add_data_validation(status_validation)
        status_validation.add("V2:V1048576")
    
    # Prepare comprehensive row data
    transcript_text = video_data.get('transcript', '')
    word_count = len(transcript_text.split()) if transcript_text else 0
    
    resolution = ""
    if video_data.get('width') and video_data.get('height'):
        resolution = f"{video_data['width']}x{video_data['height']}"
    
    row_data = [
        video_data.get('index', ''),
        video_data.get('generated_name', ''),
        video_data.get('title', ''),
        video_data.get('description', ''),  # Include full description
        video_data.get('date_processed', ''),
        video_data.get('username', ''),
        video_data.get('uploader_id', ''),
        video_data.get('channel_id', ''),
        video_data.get('channel_url', ''),
        video_data.get('video_id', ''),
        video_data.get('platform', ''),
        video_data.get('duration', 0),
        resolution,
        video_data.get('view_count', ''),
        video_data.get('like_count', ''),
        video_data.get('comment_count', ''),
        video_data.get('upload_date', ''),
        f"{video_data.get('filesize', 0) / (1024*1024):.2f}",
        video_data.get('video_path', ''),
        video_data.get('thumbnail_path', ''),
        video_data.get('transcript_path', ''),
        video_data.get('audio_path', ''),
        transcript_text,  # Include full transcript
        word_count,
        video_data.get('webpage_url', ''),
        video_data.get('status', 'In Progress'),
        f"{video_data.get('processing_time', 0):.1f}",
        video_data.get('notes', ''),
        video_data.get('error_details', '')
    ]
    
    ws.append(row_data)
    
    # Apply validation to new row
    new_row = ws.max_row
    status_validation.add(f"V{new_row}")
    
    wb.save(excel_path)

def get_drive_service():
    """Authenticate with Google Drive with detailed logging"""
    try:
        logger.log_step(0, "GDRIVE_AUTH_START", {"message": "Starting Google Drive authentication"})
        
        creds = None
        if os.path.exists(TOKEN_FILE):
            try:
                creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
                logger.log_step(0, "TOKEN_LOADED", {"message": f"Loaded existing token from {TOKEN_FILE}"})
            except Exception as e:
                logger.log_error(0, f"Failed to load token: {str(e)}")
        
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    creds.refresh(Request())
                    with open(TOKEN_FILE, 'w') as token:
                        token.write(creds.to_json())
                    logger.log_step(0, "TOKEN_REFRESHED", {"message": "Successfully refreshed token"})
                except RefreshError as e:
                    logger.log_error(0, f"Token refresh failed: {str(e)}")
                    creds = None
            
            if not creds:
                logger.log_step(0, "OAUTH_FLOW_START", {"message": "Starting OAuth flow"})
                flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
                creds = flow.run_local_server(port=8080)
                with open(TOKEN_FILE, 'w') as token:
                    token.write(creds.to_json())
                logger.log_step(0, "OAUTH_COMPLETE", {"message": "OAuth flow completed, token saved"})
        
        service = build('drive', 'v3', credentials=creds)
        logger.log_step(0, "GDRIVE_SERVICE_READY", {"message": "Google Drive service authenticated"})
        return service
        
    except Exception as e:
        logger.log_error(0, f"Google Drive authentication failed: {str(e)}")
        raise

def upload_to_drive(file_path: str):
    """Upload Excel file to Google Drive with detailed logging"""
    try:
        upload_start = time.time()
        file_size = os.path.getsize(file_path)
        
        logger.log_step(0, "GDRIVE_UPLOAD_START", {
            "message": f"Uploading to Google Drive",
            "filename": os.path.basename(file_path),
            "size_mb": round(file_size / (1024*1024), 2)
        })
        
        service = get_drive_service()
        
        # Create or find folder
        folder_query = f"name='{DRIVE_FOLDER}' and mimeType='application/vnd.google-apps.folder'"
        results = service.files().list(q=folder_query).execute()
        folders = results.get('files', [])
        
        if folders:
            folder_id = folders[0]['id']
            logger.log_step(0, "GDRIVE_FOLDER_FOUND", {"message": f"Found existing folder: {DRIVE_FOLDER}"})
        else:
            folder_metadata = {'name': DRIVE_FOLDER, 'mimeType': 'application/vnd.google-apps.folder'}
            folder = service.files().create(body=folder_metadata).execute()
            folder_id = folder['id']
            logger.log_step(0, "GDRIVE_FOLDER_CREATED", {"message": f"Created new folder: {DRIVE_FOLDER}"})
        
        # Prepare upload
        filename = os.path.basename(file_path)
        media = MediaFileUpload(file_path, resumable=True)
        
        # Check if file exists
        file_query = f"name='{filename}' and '{folder_id}' in parents"
        existing = service.files().list(q=file_query).execute().get('files', [])
        
        if existing:
            # Update existing file
            service.files().update(fileId=existing[0]['id'], media_body=media).execute()
            upload_time = time.time() - upload_start
            logger.log_step(0, "GDRIVE_UPLOAD_UPDATE", {
                "message": f"Updated existing file: {filename}",
                "upload_time": round(upload_time, 2)
            })
        else:
            # Create new file
            file_metadata = {'name': filename, 'parents': [folder_id]}
            service.files().create(body=file_metadata, media_body=media).execute()
            upload_time = time.time() - upload_start
            logger.log_step(0, "GDRIVE_UPLOAD_CREATE", {
                "message": f"Created new file: {filename}",
                "upload_time": round(upload_time, 2)
            })
        
        logger.log_file_operation("GDRIVE_UPLOAD", file_path, f"Google Drive: {DRIVE_FOLDER}/{filename}", True, file_size)
            
    except Exception as e:
        logger.log_error(0, f"Google Drive upload failed: {str(e)}", {"file_path": file_path})

# ---------------------------
# Main Processing Pipeline
# ---------------------------
def process_single_video(url: str, index: int = 1) -> Dict[str, Any]:
    """Process a single video through the complete pipeline"""
    start_time = time.time()
    date_processed = datetime.now().strftime('%d.%m.%Y')
    
    result = {
        'index': index,
        'source_url': url,
        'date_processed': date_processed,
        'status': 'Failed',
        'processing_time': 0,
        'notes': '',
        'error_details': ''
    }
    
    try:
        logger.log_step(index, "PROCESSING_START", {"message": f"Starting complete pipeline", "url": url})
        
        # Step 1: Download video and extract metadata
        video_path, metadata, raw_info = download_video_and_metadata(url, index)
        result.update(metadata)
        result['video_path'] = video_path
        
        # Step 2: Download thumbnail (parallel to other operations)
        thumbnail_path = download_thumbnail(
            metadata.get('thumbnail_url'), 
            metadata.get('video_id'), 
            metadata.get('username'),
            index
        )
        result['thumbnail_path'] = thumbnail_path
        
        # Step 3: Generate smart video name
        generated_name = generate_smart_video_name(
            metadata.get('title', ''), 
            metadata.get('description', ''),
            index
        )
        result['generated_name'] = generated_name
        
        # Step 4: Convert to audio
        audio_path = convert_video_to_audio(video_path, index)
        result['audio_path'] = audio_path
        
        # Step 5: Transcribe
        transcript = transcribe_audio_with_whisper(audio_path, index)
        result['transcript'] = transcript
        
        # Step 6: Save transcript as separate file
        if transcript:
            transcript_path = save_transcript_file(transcript, generated_name, metadata, index)
            result['transcript_path'] = transcript_path
            result['status'] = 'Completed'
            result['notes'] = f'Successfully processed. Transcript: {len(transcript)} chars, {len(transcript.split())} words'
            
            logger.log_step(index, "PIPELINE_SUCCESS", {
                "message": f"Complete pipeline successful: {generated_name}",
                "transcript_words": len(transcript.split()),
                "files_created": [
                    os.path.basename(video_path),
                    os.path.basename(thumbnail_path) if thumbnail_path else None,
                    os.path.basename(transcript_path) if transcript_path else None
                ]
            })
        else:
            result['status'] = 'Needs Review'
            result['notes'] = 'Video processed but transcription failed'
            result['error_details'] = 'Empty transcription result'
            
            logger.log_step(index, "PIPELINE_PARTIAL", {
                "message": "Video downloaded but transcription failed"
            })
        
        # Handle audio file based on configuration
        if KEEP_AUDIO_FILES:
            logger.log_step(index, "AUDIO_PRESERVED", {
                "message": f"Audio file preserved: {os.path.basename(audio_path)}",
                "path": audio_path,
                "size_mb": round(os.path.getsize(audio_path) / (1024*1024), 2) if os.path.exists(audio_path) else 0
            })
            result['audio_path'] = audio_path # Keep in result if preserving
        else:
            try:
                if os.path.exists(audio_path):
                    os.remove(audio_path)
                    logger.log_file_operation("AUDIO_CLEANUP", audio_path, "", True, 0)
                    logger.log_step(index, "AUDIO_CLEANED", {"message": "Temporary audio file removed"})
                result['audio_path'] = "" # Clear from result if removed
            except Exception as e:
                logger.log_error(index, f"Failed to clean up audio file: {str(e)}")
                result['audio_path'] = audio_path # Keep reference if cleanup failed
        
    except Exception as e:
        error_msg = str(e)
        result['status'] = 'Failed'
        result['notes'] = f'Processing failed: {error_msg}'
        result['error_details'] = error_msg
        logger.log_error(index, f"Pipeline failed: {error_msg}", {"url": url})
    
    finally:
        result['processing_time'] = time.time() - start_time
        logger.log_step(index, "PROCESSING_COMPLETE", {
            "message": f"Processing finished",
            "status": result['status'],
            "total_time": round(result['processing_time'], 2)
        })
    
    return result

def process_batch(urls: List[str]) -> List[Dict[str, Any]]:
    """Process multiple videos with comprehensive logging"""
    logger.log_step(0, "BATCH_START", {
        "message": f"Starting batch processing",
        "total_urls": len(urls),
        "session_id": logger.session_id
    })
    
    results = []
    
    for i, url in enumerate(urls, 1):
        result = process_single_video(url.strip(), i)
        results.append(result)
        
        # Update Excel after each video
        update_excel_comprehensive(result, EXCEL_FILE_PATH)
        
        # Log progress
        logger.log_step(0, "BATCH_PROGRESS", {
            "message": f"Progress: {i}/{len(urls)} completed",
            "current_status": result['status']
        })
    
    # Finalize session logging
    logger.finalize_session(results)
    
    # Upload to Google Drive
    upload_to_drive(EXCEL_FILE_PATH)
    
    return results

# ---------------------------
# Main Entry Point
# ---------------------------
def main():
    parser = argparse.ArgumentParser(description='Enhanced Video Downloader with Transcription & Intensive Logging')
    parser.add_argument('--url', help='Single video URL to process')
    parser.add_argument('--urls-file', help='Text file with one URL per line')
    parser.add_argument('--batch-size', type=int, default=10, help='Process videos in batches (for future use)')
    
    args = parser.parse_args()
    
    urls = []
    
    if args.url:
        urls.append(args.url.strip())
    elif args.urls_file:
        try:
            with open(args.urls_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#'):
                        urls.append(line)
        except FileNotFoundError:
            print(f"URLs file not found: {args.urls_file}")
            return
    else:
        # Interactive mode
        url = input("Enter video URL (Instagram, TikTok, YouTube, etc.): ").strip()
        if url:
            urls.append(url)
        else:
            print("No URL provided.")
            return
    
    if not urls:
        print("No valid URLs found.")
        return
    
    print(f"\nStarting Enhanced Video Downloader")
    print(f"Session ID: {logger.session_id}")
    print(f"Videos to process: {len(urls)}")
    print(f"Directories:")
    print(f" 📹 Videos: {VIDEO_OUTPUT_DIR}")
    print(f" 🎵 Audio: {AUDIO_OUTPUT_DIR}")
    print(f" 🖼️ Thumbnails: {THUMBNAILS_DIR}")
    print(f" 📄 Transcripts: {TRANSCRIPTS_DIR}")
    print(f" 📊 Excel: {EXCEL_FILE_PATH}")
    print(f" 📋 Logs: {LOGS_DIR}")
    
    # Process videos
    results = process_batch(urls)
    
    print(f"\nProcessing complete! Check logs in {LOGS_DIR} for detailed session information.")

if __name__ == '__main__':
    main()