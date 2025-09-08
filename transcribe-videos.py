import os
import re
import yt_dlp  # For downloading videos from Instagram/TikTok
import ffmpeg  # For converting video to audio
import whisper  # For transcribing audio
from openai import OpenAI  # Client-based API for OpenAI v1+
from datetime import datetime
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.exceptions import InvalidFileException
from dotenv import load_dotenv  # For loading .env files

# Google API imports for Drive upload
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.auth.exceptions import RefreshError
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# ---------------------------
# Load environment variables (if .env exists)
# ---------------------------
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception as e:
    print(f"Note: .env file not loaded ({str(e)}). Will use defaults or environment variables.")

# ---------------------------
# Configuration Variables (from .env with safe fallbacks)
# ---------------------------
# These can be set in a .env file or environment. If absent, sensible defaults are used.
VIDEO_OUTPUT_DIR = os.getenv("VIDEO_OUTPUT_DIR", "videos")
TRANSCRIPTS_DIR = os.getenv("TRANSCRIPTS_DIR", "transcripts")
EXCEL_FILENAME = os.getenv("EXCEL_FILENAME", "transcripts.xlsx")
# Allow full override of the excel path, otherwise build from TRANSCRIPTS_DIR + EXCEL_FILENAME
EXCEL_FILE_PATH = os.getenv("EXCEL_FILE_PATH", os.path.join(TRANSCRIPTS_DIR, EXCEL_FILENAME))

# Google Drive configuration
SCOPES = ['https://www.googleapis.com/auth/drive.file']
CREDENTIALS_FILE = "credentials.json"  # Your Google API credentials file
TOKEN_FILE = "token.json"              # Will store OAuth tokens
DRIVE_FOLDER = "AIWaverider"           # Folder on Drive where the file will be placed

# ---------------------------
# OpenAI Configuration
# ---------------------------
# Ensure OPENAI_API_KEY is set in your .env or environment
api_key = os.getenv("OPENAI_API_KEY")
if not api_key:
    raise ValueError("OPENAI_API_KEY not set. Please add it to your .env or environment variables.")
client = OpenAI(api_key=api_key)

# ---------------------------
# Utility Functions
# ---------------------------
def get_unique_filename(path):
    """
    If path exists, append a counter to make it unique.
    """
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    counter = 1
    while True:
        new_path = f"{base}_{counter}{ext}"
        if not os.path.exists(new_path):
            return new_path
        counter += 1

# ---------------------------
# Video & Transcript Pipeline
# ---------------------------
def get_video_number(output_dir, username):
    """
    Get the next available number for a username by scanning existing files.
    Returns the next number to use (1-based).
    """
    pattern = re.compile(rf'\d+_{re.escape(username)}_.*\.mp4$')
    max_num = 0
    
    if os.path.exists(output_dir):
        for filename in os.listdir(output_dir):
            if pattern.match(filename):
                try:
                    num = int(filename.split('_')[0])
                    max_num = max(max_num, num)
                except (ValueError, IndexError):
                    continue
    return max_num + 1

def is_video_downloaded(output_dir, video_id):
    """
    Check if a video with this ID already exists in the output directory.
    """
    if os.path.exists(output_dir):
        for filename in os.listdir(output_dir):
            if f"_{video_id}." in filename:
                return True
    return False

def download_video(url, output_dir):
    """
    Download the video from the URL and save it with format: NN_username_id.mp4
    Returns the file path, metadata, and info dict.
    Skips download if video ID already exists in output_dir.
    """
    os.makedirs(output_dir, exist_ok=True)

    # First extract info without downloading
    with yt_dlp.YoutubeDL({'format': 'mp4'}) as ydl:
        info = ydl.extract_info(url, download=False)
        video_id = info['id']
        username = info.get('uploader', 'unknown').lower()
        username = re.sub(r'[^\w]+', '_', username)  # sanitize username
        
        # Check if video already exists
        if is_video_downloaded(output_dir, video_id):
            print(f"Video {video_id} already exists in {output_dir}, skipping download.")
            # Find the existing file
            for filename in os.listdir(output_dir):
                if f"_{video_id}." in filename:
                    return os.path.join(output_dir, filename), {'duration': info.get('duration', 0)}, info

        # Get next number for this username
        num = get_video_number(output_dir, username)
        # Create filename: 01_username_id.mp4
        template = os.path.join(output_dir, f"{num:02d}_{username}_{video_id}.%(ext)s")
        ydl_opts = {'outtmpl': template, 'format': 'mp4'}

    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        info = ydl.extract_info(url, download=True)
        original_file = ydl.prepare_filename(info)

    metadata = {
        'duration': info.get('duration', 0),
        'filesize': os.path.getsize(original_file),
    }
    return original_file, metadata, info


def convert_video_to_audio(video_file, audio_file):
    """
    Extract high-quality audio from video file
    """
    ffmpeg.input(video_file).output(
        audio_file,
        acodec='pcm_s16le',    # Use high-quality codec
        ac=1,                  # Mono channel
        ar='16k',             # 16kHz sample rate (recommended for Whisper)
        loglevel='error'      # Show errors only
    ).run(overwrite_output=True)


def chunk_audio(audio_file, chunk_duration=30):
    """
    Split long audio into manageable chunks using pure Python (no pydub required)
    """
    import wave
    import math
    
    # Open the WAV file
    with wave.open(audio_file, 'rb') as wav:
        # Get the WAV file properties
        n_channels = wav.getnchannels()
        sampwidth = wav.getsampwidth()
        framerate = wav.getframerate()
        n_frames = wav.getnframes()
        
        # Calculate frames per chunk based on duration
        frames_per_chunk = int(framerate * chunk_duration)
        n_chunks = math.ceil(n_frames / frames_per_chunk)
        chunks = []
        
        for i in range(n_chunks):
            chunk_path = f"{audio_file}_chunk_{i}.wav"
            # Create a new WAV file for this chunk
            with wave.open(chunk_path, 'wb') as chunk_wav:
                chunk_wav.setnchannels(n_channels)
                chunk_wav.setsampwidth(sampwidth)
                chunk_wav.setframerate(framerate)
                
                # Read frames for this chunk
                start_frame = i * frames_per_chunk
                wav.setpos(start_frame)
                frames_to_read = min(frames_per_chunk, n_frames - start_frame)
                chunk_frames = wav.readframes(frames_to_read)
                
                # Write frames to chunk file
                chunk_wav.writeframes(chunk_frames)
                chunks.append(chunk_path)
    
    return chunks


def transcribe_audio(audio_file):
    """
    Transcribe audio with memory management
    """
    import torch
    import gc
    
    try:
        # Clear CUDA cache if using GPU
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
        
        model = whisper.load_model('base', device='cuda' if torch.cuda.is_available() else 'cpu') # small, medium, large
        result = model.transcribe(audio_file, verbose=True)
        transcript = result['text']
        
        # Cleanup
        del model
        gc.collect()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            
        return transcript
        
    except Exception as e:
        print(f"Transcription error: {str(e)}")
        return ""


def transcribe_long_audio(audio_file):
    """
    Handle long audio files by chunking
    """
    chunks = chunk_audio(audio_file)
    full_transcript = []
    
    for chunk in chunks:
        transcript = transcribe_audio(chunk)
        full_transcript.append(transcript)
        os.remove(chunk)  # Clean up chunk file
    
    return ' '.join(full_transcript)


def generate_video_name(caption):
    """
    Use GPT-4o-mini to extract or generate a one-word name from the caption,
    then append today's date in dd.MM.YYYY format.
    """
    date_str = datetime.now().strftime('%d.%m.%Y')
    prompt = (
        f"Extract the tool name from this caption: \"{caption}\". "
        "If no tool name is present, create a one-word name summarizing the content. "
        "Return only that single word."
    )
    response = client.chat.completions.create(
        model='gpt-4o-mini',
        messages=[
            {'role': 'system', 'content': 'You extract or generate a one-word tool name.'},
            {'role': 'user', 'content': prompt}
        ]
    )
    name = response.choices[0].message.content.strip().split()[0].lower()
    return f"{name}_{date_str}"


def update_excel(video_name, description, transcript, date_str, metadata, source_url, excel_path):
    """
    Append a row to Excel with columns:
    Video Name | Description | Transcript | Date | Size | Duration | Source URL | Status
    Default Status = "In Progress" with dropdown list of ["In Progress","Completed"].
    """
    excel_dir = os.path.dirname(excel_path)
    # If excel_path is just a filename (no dir), os.path.dirname returns ''.
    if excel_dir:
        os.makedirs(excel_dir, exist_ok=True)

    # Set up data validation for Status column
    dv = DataValidation(type="list", formula1='"In Progress,Completed"', allow_blank=False)

    # Load or create workbook
    if os.path.exists(excel_path):
        try:
            # Try to open existing workbook
            wb = load_workbook(excel_path)
            ws = wb.active
        except InvalidFileException:
            # CHANGE: handle invalid/corrupt excel file by moving it aside and creating a fresh workbook
            # This prevents openpyxl.InvalidFileException from crashing the run.
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            bad_path = f"{excel_path}.invalid_{timestamp}"
            try:
                os.rename(excel_path, bad_path)
                print(f"Moved invalid Excel file '{excel_path}' to '{bad_path}' and will create a new workbook.")
            except Exception as e:
                print(f"Failed to move invalid Excel file: {e}. Attempting to continue by creating a new workbook.")
            wb = Workbook()
            ws = wb.active
            # Header row for new workbook
            ws.append(["Video Name", "Description", "Transcript", "Date", "Size", "Duration", "Source URL", "Status"])
            ws.add_data_validation(dv)
            dv.add("H2:H1048576")
    else:
        wb = Workbook()
        ws = wb.active
        # Header row
        ws.append(["Video Name", "Description", "Transcript", "Date", "Size", "Duration", "Source URL", "Status"])
        ws.add_data_validation(dv)
        # Apply dropdown to entire Status column (rows 2 to infinity)
        dv.add("H2:H1048576")  # Corrected line: uses add() instead of ranges.append()

    # Prepare row data
    size_str = f"{metadata['filesize'] / (1024*1024):.2f} MB"
    duration_str = f"{int(metadata['duration'])} Seconds"
    row = [video_name, description, transcript, date_str, size_str, duration_str, source_url, "In Progress"]

    # Append row and apply validation to the new cell
    ws.append(row)
    new_row = ws.max_row
    dv.add(f"H{new_row}")  # Already correct

    wb.save(excel_path)
# ---------------------------
# Google Drive Upload Functions
# ---------------------------
def get_drive_service():
    """
    Authenticate with Google Drive using OAuth, refreshing if possible.
    Falls back to full flow on refresh failure.
    """
    creds = None
    if os.path.exists(TOKEN_FILE):
        try:
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        except Exception as e:
            print(f"Failed to load token file '{TOKEN_FILE}': {e}. Will perform full OAuth flow.")
            creds = None

    # If no valid credentials, try to refresh; otherwise start full OAuth flow.
    if not creds or not creds.valid:
        # If credentials exist but are expired, try to refresh and overwrite token file
        if creds and creds.expired:
            try:
                creds.refresh(Request())
                # Save refreshed token
                with open(TOKEN_FILE, 'w') as token:
                    token.write(creds.to_json())
            except RefreshError as e:
                print(f"Token refresh failed: {e}. Removing stale token and starting new OAuth flow.")
                try:
                    os.remove(TOKEN_FILE)
                except Exception:
                    pass
                creds = None
            except Exception as e:
                print(f"Unexpected error during token refresh: {e}. Attempting full OAuth flow.")
                creds = None

        # If credentials are still missing or invalid, run full OAuth flow
        if not creds:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=8080, access_type='offline', prompt='consent')
            # Save new token
            with open(TOKEN_FILE, 'w') as token:
                token.write(creds.to_json())

    return build('drive', 'v3', credentials=creds)


def get_folder_id(service):
    query = f"name='{DRIVE_FOLDER}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
    files = results.get('files', [])
    if files:
        return files[0]['id']
    folder_metadata = {'name': DRIVE_FOLDER, 'mimeType': 'application/vnd.google-apps.folder'}
    folder = service.files().create(body=folder_metadata, fields='id').execute()
    return folder['id']


def upload_excel_to_drive(file_path):
    service = get_drive_service()
    folder_id = get_folder_id(service)
    gs_name = os.path.basename(file_path).rsplit('.', 1)[0]

    query = (
        f"name='{gs_name}' and '{folder_id}' in parents and trashed=false "
        f"and mimeType='application/vnd.google-apps.spreadsheet'"
    )
    results = service.files().list(q=query, spaces='drive', fields='files(id)').execute()
    files = results.get('files', [])

    media = MediaFileUpload(
        file_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True
    )

    if files:
        service.files().update(fileId=files[0]['id'], media_body=media).execute()
    else:
        file_metadata = {'name': gs_name, 'parents': [folder_id], 'mimeType': 'application/vnd.google-apps.spreadsheet'}
        service.files().create(body=file_metadata, media_body=media, fields='id').execute()

# ---------------------------
# Main Process
# ---------------------------
def main():
    parser = argparse.ArgumentParser(description='Transcribe videos from URL(s) and update Excel.')
    parser.add_argument('--url', help='Single video URL to process')
    parser.add_argument('--urls-file', help='Path to a text file containing one URL per line')
    args = parser.parse_args()

    urls = []
    if args.url:
        urls.append(args.url.strip())
    if args.urls_file:
        if not os.path.exists(args.urls_file):
            print(f"URLs file not found: {args.urls_file}")
            return
        with open(args.urls_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:
                    urls.append(line)

    # If no args provided, fall back to interactive single URL input
    if not urls:
        single = input("Enter video URL (Instagram or TikTok): ").strip()
        if not single:
            print("No URL provided. Exiting.")
            return
        urls.append(single)

    # Process each URL and append to Excel. Upload once at the end.
    date_str = datetime.now().strftime('%d.%m.%Y')
    for url in urls:
        print(f"Processing: {url}")
        print("Downloading video...")
        video_file, metadata, info = download_video(url, VIDEO_OUTPUT_DIR)
        print(f"Downloaded video to: {video_file}")

        caption = info.get('description', '').strip()

        print("Generating video name via GPT-4o-mini...")
        final_name = generate_video_name(caption)

        ext = os.path.splitext(video_file)[1]
        new_video_path = os.path.join(VIDEO_OUTPUT_DIR, final_name + ext)
        new_video_path = get_unique_filename(new_video_path)
        os.rename(video_file, new_video_path)
        video_file = new_video_path
        metadata['filesize'] = os.path.getsize(video_file)

        audio_file = os.path.splitext(video_file)[0] + '.wav'
        print("Converting video to audio...")
        convert_video_to_audio(video_file, audio_file)

        print("Transcribing audio...")
        if metadata['duration'] > 30:
            transcript = transcribe_long_audio(audio_file)
        else:
            transcript = transcribe_audio(audio_file)

        if not transcript:
            print("Warning: Transcription failed or produced empty result")
        else:
            print(f"Transcription length: {len(transcript)} characters")

        print(f"Video Name: {final_name}")
        print("--- Transcript ---")
        print(transcript)
        print("------------------")

        if os.path.exists(audio_file):
            os.remove(audio_file)

        update_excel(final_name, caption, transcript, date_str, metadata, url, EXCEL_FILE_PATH)
        print(f"Appended record to Excel: {EXCEL_FILE_PATH}")

    # After processing all URLs, upload the Excel once
    print("Uploading Excel to Google Drive...")
    upload_excel_to_drive(EXCEL_FILE_PATH)
    print("Upload complete.")

if __name__ == '__main__':
    main()
